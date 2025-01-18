from openpyxl import load_workbook
from datetime import datetime as dt

arquivo = load_workbook('C:\\Users\\erirodrigues\\Documents\\Estudo.xlsx')

qtd_semana_retrasada = float(input("Digite a quantidade de transações semana retrasada XXXXXX: "))
tpv_semana_retrasada = float(input('Digite o TPV de semana retrasada XXXXXXXX.XX: '))
data_semana_retrasada = input('Digite a data da semana retrasada (AAAA-MM-DD): ')

qtd_semana_passada = float(input("Digite a quantidade de transações semana passada XXXXXX: "))
tpv_semana_passada = float(input('Digite o TPV de semana passada XXXXXXXX.XX: '))
data_semana_passada = input('Digite a data da semana passada (AAAA-MM-DD): ')

qtd_hj = float(input("Digite a quantidade de transações no dia da crise XXXXXX: "))
tpv_hj = float(input('Digite o TPV do dia da crise XXXXXXXX.XX: '))
data_hj = input('Digite a data do dia da crise (AAAA-MM-DD): ')

hora_inicio = input('Digite a hora de início (hh:mm): ')
hora_final = input('Digite a hora final (hh:mm): ')

# ver as abas
# print(arquivo.sheetnames)

# aba ativa
aba_atual = arquivo.active
print(aba_atual)

# selecionar uma aba específica
aba_teste = arquivo['Planilha1']
# print(aba_teste)

# selecionar células
# print(aba_teste['C4'].value)

# inserindo valores de quantidade
aba_teste.cell(row = 4, column = 3).value = qtd_semana_retrasada
aba_teste.cell(row = 5, column = 3).value = qtd_semana_passada
aba_teste.cell(row = 6, column = 3).value = qtd_hj

# inserindo valores de TPV
aba_teste.cell(row = 4, column = 4).value = tpv_semana_retrasada
aba_teste.cell(row = 5, column = 4).value = tpv_semana_passada
aba_teste.cell(row = 6, column = 4).value = tpv_hj

# inserindo hora inicial e final
aba_teste.cell(row = 4, column = 5).value = f'{hora_inicio} até {hora_final}'
aba_teste.cell(row = 5, column = 5).value = f'{hora_inicio} até {hora_final}'
aba_teste.cell(row = 6, column = 5).value = f'{hora_inicio} até {hora_final}'

# inserindo a data dos eventos
dias_semanapt = [
    "Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sabado", "Domingo"
]

dataR = dt.strptime(data_semana_retrasada, '%Y-%m-%d') # convertendo para objeto datetime
dia_da_semanaR = dias_semanapt[dataR.weekday()] # Obtendo o dia da semana em português
data_semana_retrasada_formatada = dataR.strftime('%d/%m/%Y')

dataP = dt.strptime(data_semana_passada, '%Y-%m-%d') # convertendo para objeto datetime
dia_da_semanaP = dias_semanapt[dataP.weekday()]  # Obtendo o dia da semana em português
data_semana_passada_formatada = dataP.strftime('%d/%m/%Y')

dataH = dt.strptime(data_hj, '%Y-%m-%d') # convertendo para objeto datetime
dia_da_semanaH = dias_semanapt[dataH.weekday()]  # Obtendo o dia da semana em português
data_hj_formatada = dataH.strftime('%d/%m/%Y')

aba_teste.cell(row = 4, column = 2).value = f'{dia_da_semanaR} <> {data_semana_retrasada_formatada}'
aba_teste.cell(row = 5, column = 2).value = f'{dia_da_semanaP} <> {data_semana_passada_formatada}'
aba_teste.cell(row = 6, column = 2).value = f'{dia_da_semanaH} <> {data_hj_formatada}'

# editar valor da célula
# aba_teste['C4'].value = 23666

# quantidade hj x sp
if qtd_hj - qtd_semana_passada > 0:
    aba_teste.cell(row = 9, column = 2).value = 'Aumento de TRS'
    aba_teste.cell(row = 9, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaP} {data_semana_passada_formatada}'
    
else:
    aba_teste.cell(row = 9, column = 2).value = 'Diminuição TRS'
    aba_teste.cell(row = 9, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaP} {data_semana_passada_formatada}'

# quantidade hj x sr    
if qtd_hj - qtd_semana_retrasada > 0:
    aba_teste.cell(row = 10, column = 2).value = 'Aumento de TRS'
    aba_teste.cell(row = 10, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaR} {data_semana_retrasada_formatada}'
    
else:
    aba_teste.cell(row = 10, column = 2).value = 'Diminuição TRS'
    aba_teste.cell(row = 10, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaR} {data_semana_retrasada_formatada}'

# TPV hj x sp
if tpv_hj - tpv_semana_passada > 0:
    aba_teste.cell(row = 36, column = 2).value = 'Aumento de TPV'
    aba_teste.cell(row = 36, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaP} {data_semana_passada_formatada}'
    
else:
    aba_teste.cell(row = 36, column = 2).value = 'Diminuição TPV'
    aba_teste.cell(row = 36, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaP} {data_semana_passada_formatada}'
    
# TPV hj x sr
if tpv_hj - tpv_semana_retrasada > 0:
    aba_teste.cell(row = 37, column = 2).value = 'Aumento de TPV'
    aba_teste.cell(row = 37, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaR} {data_semana_retrasada_formatada}'
    
else:
    aba_teste.cell(row = 37, column = 2).value = 'Diminuição TPV'
    aba_teste.cell(row = 37, column = 4).value = f'entre  {dia_da_semanaH} {data_hj_formatada} comparando com {dia_da_semanaR} {data_semana_retrasada_formatada}'

arquivo.save('Estudos2.xlsx')

#ultima linha
# print(aba_teste.max_row)
# tamanho da coluna A
# print(len(aba_teste['A']))